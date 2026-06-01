#!/usr/bin/env python3
"""
L2 — Profiling de la source Satisfaction (e-Satis / IQSS, fichiers plats)
Tâche : [P3] Profiling + mapping + cleaning Satisfaction (869dh1d8b)

Source la plus "salissante" du projet : fichiers plats à encodage/séparateur
variables. Ce script détecte, par fichier : encoding, séparateur, présence
d'en-tête, nb de lignes, puis pour les colonnes clés (finess, region,
score_all_rea_ajust) : % de nulls, cardinalité, distribution du score et
anomalies (hors plage 0-100, mojibake d'encodage).

Usage :
    python3 scripts/satisfaction_profiling.py [REP_SATISFACTION]
    # défaut : ./data/raw/satisfaction  (adapter au montage local des sources)

Dépendances optionnelles : chardet (détection d'encodage), openpyxl (.xlsx).
Sans elles, le script bascule sur une heuristique utf-8 -> latin-1 et ignore
les .xlsx (en le signalant).
"""
import csv, io, os, sys, glob, statistics as st

KEY_COLS = ["finess", "region", "score_all_rea_ajust"]


def detect_encoding(path):
    """chardet si dispo, sinon heuristique : utf-8 strict, sinon latin-1."""
    raw = open(path, "rb").read(200_000)
    try:
        import chardet
        guess = chardet.detect(raw)
        enc = (guess.get("encoding") or "latin-1")
        # chardet renvoie souvent 'ISO-8859-1' / 'Windows-1252' pour e-Satis
        return enc, round(guess.get("confidence", 0), 2)
    except ImportError:
        try:
            raw.decode("utf-8")
            return "utf-8", None
        except UnicodeDecodeError:
            return "latin-1 (heuristique)", None


def sniff_sep(path, enc):
    line = io.open(path, encoding=enc, errors="replace").readline()
    return max([";", ",", "\t", "|"], key=lambda s: line.count(s))


def col_index(header, name):
    """Match exact d'abord (évite de confondre `score_all_rea_ajust` avec la
    colonne de comptage `nb_rep_score_all_rea_ajust`), puis repli sur un
    'contient' qui exclut les colonnes de dénombrement (`nb_...`)."""
    norm = [str(c).strip().lower() for c in header]
    if name in norm:
        return norm.index(name)
    for i, c in enumerate(norm):
        if name in c and not c.startswith("nb"):
            return i
    return None


def read_rows(path, enc):
    """Renvoie (header, rows) pour csv ou xlsx."""
    if path.lower().endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = [str(c) if c is not None else "" for c in next(it)]
        rows = [list(r) for r in it]
        return header, rows, ";(xlsx)"
    sep = sniff_sep(path, enc)
    with io.open(path, encoding=enc, errors="replace", newline="") as fh:
        r = csv.reader(fh, delimiter=sep)
        header = next(r)
        rows = [row for row in r if any((c or "").strip() for c in row)]
    return header, rows, sep


def to_float(x):
    if x is None:
        return None
    s = str(x).replace(",", ".").strip()
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def profile_file(path):
    enc, conf = detect_encoding(path)
    name = os.path.relpath(path)
    print("\n" + "=" * 78)
    print(f"FICHIER : {name}")
    print(f"  encoding   : {enc}" + (f" (confiance {conf})" if conf else ""))
    try:
        header, rows, sep = read_rows(path, enc if "heuristique" not in enc else "latin-1")
    except ImportError as e:
        print(f"  [SKIP] dépendance manquante pour ce format : {e}")
        return
    print(f"  séparateur : {sep!r}")
    print(f"  en-tête    : oui ({len(header)} colonnes)")
    print(f"  lignes     : {len(rows)}")

    idx = {c: col_index(header, c) for c in KEY_COLS}
    print(f"  colonnes clés détectées : "
          + ", ".join(f"{c}->#{idx[c]}" for c in KEY_COLS))

    for c in KEY_COLS:
        i = idx[c]
        if i is None:
            print(f"   - {c:22} : ABSENTE")
            continue
        vals = [(r[i] if i < len(r) else None) for r in rows]
        nn = [v for v in vals if v not in (None, "")]
        nulls = len(vals) - len(nn)
        pct = round(100 * nulls / len(vals), 1) if vals else 0
        line = f"   - {c:22} : nulls {nulls} ({pct}%), distinct {len(set(map(str, nn)))}"
        if c == "score_all_rea_ajust":
            f = [to_float(v) for v in nn]
            f = [x for x in f if x is not None]
            if f:
                anomalies = sum(1 for x in f if x < 0 or x > 100)
                line += (f" | score n={len(f)} min={min(f):.1f} max={max(f):.1f} "
                         f"moy={st.mean(f):.1f} | hors[0-100]={anomalies}")
        if c == "region":
            mojibake = sum(1 for v in nn if "�" in str(v) or "Ã" in str(v))
            line += f" | mojibake={mojibake}"
        print(line)


def main():
    base = sys.argv[1] if len(sys.argv) > 1 else "data/raw/satisfaction"
    if not os.path.isdir(base):
        print(f"[!] Répertoire introuvable : {base}")
        print("    Passe le chemin du montage local des fichiers Satisfaction en argument.")
        sys.exit(1)
    files = []
    for f in glob.glob(os.path.join(base, "**", "*"), recursive=True):
        b = os.path.basename(f).lower()
        if (b.endswith((".csv", ".xlsx")) and not b.startswith("~$")
                and "esatis48" in b and ("donnee" in b or "resultat" in b)):
            files.append(f)
    files = sorted(set(files))
    if not files:
        print(f"[!] Aucun fichier e-Satis 48h trouvé sous {base}")
        sys.exit(1)
    print(f"PROFILING SATISFACTION — {len(files)} fichier(s) e-Satis 48h MCO")
    for f in files:
        profile_file(f)


if __name__ == "__main__":
    main()
