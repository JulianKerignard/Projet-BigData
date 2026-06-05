#!/usr/bin/env python3
"""Dashboard Satisfaction — style Power BI (mode custom).

Besoin imposé B8 : taux global de satisfaction par région, campagne 2020.
Particularité : la mesure est un SCORE MOYEN (pas un compteur) → le moteur de
cross-filter sum-based ne s'applique pas. On utilise le mode `custom_script` de
dashboard_common (thème + helpers partagés, rendu sur mesure) en réutilisant le
nouveau style (cartes KPI en élévation, icônes, pastilles delta, #scope, data
labels de valeur, dégradés cohérents).

Reproductible OFFLINE : les agrégats sont lus depuis viz/data_satisfaction.json
(commité, aucune donnée nominative). Le xlsx source (gitignoré) n'est lu QUE pour
(re)générer ce cache quand il est présent — comme extract_deces.py pour le CSV.

Source brute (optionnelle) : DATA 2024/Satisfaction/2020/resultats-esatis48h-mco-open-data-2020.xlsx
Cache  : viz/data_satisfaction.json
Sortie : viz/satisfaction_dashboard.html
"""
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import dashboard_common as dc

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "DATA 2024/Satisfaction/2020/resultats-esatis48h-mco-open-data-2020.xlsx"
CACHE = ROOT / "viz/data_satisfaction.json"
OUT = ROOT / "viz/satisfaction_dashboard.html"

# sous-dimensions de l'expérience patient (libellé -> index colonne score)
SUBSCORES = {
    "Accueil": 12, "Prise en charge infirmiers": 14, "Prise en charge médecins": 16,
    "Chambre": 18, "Repas": 20, "Organisation sortie": 22,
}

# normalisation des libellés région du fichier e-Satis vers les noms canoniques
# du GeoJSON (découpage 2016) pour que la carte matche
REGION_NORM = {
    "Ile de France": "Île-de-France", "Hauts de France": "Hauts-de-France",
    "Nouvelle Aquitaine": "Nouvelle-Aquitaine", "PACA": "Provence-Alpes-Côte d'Azur",
    "Grand-Est": "Grand Est", "Centre Val de Loire": "Centre-Val de Loire",
    "Bourgogne Franche Comté": "Bourgogne-Franche-Comté",
    "Auvergne Rhône Alpes": "Auvergne-Rhône-Alpes",
}


def norm_region(r):
    r = (r or "").strip()
    return REGION_NORM.get(r, r)


def num(x):
    if x is None:
        return None
    try:
        return float(str(x).replace(",", "."))
    except ValueError:
        return None


def extract_from_xlsx():
    """Recalcule les agrégats depuis le xlsx source (e-Satis 48h MCO 2020)."""
    import openpyxl

    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    next(it)  # header

    region_sum, region_n = {}, {}
    sub_sum = {k: 0.0 for k in SUBSCORES}
    sub_n = {k: 0 for k in SUBSCORES}
    diffuses, total = 0, 0
    for row in it:
        total += 1
        region = norm_region(row[4])
        sg = num(row[8])
        if region and sg is not None and 0 <= sg <= 100:
            diffuses += 1
            region_sum[region] = region_sum.get(region, 0.0) + sg
            region_n[region] = region_n.get(region, 0) + 1
            for k, idx in SUBSCORES.items():
                v = num(row[idx])
                if v is not None and 0 <= v <= 100:
                    sub_sum[k] += v
                    sub_n[k] += 1

    regions = sorted(region_sum, key=lambda r: region_sum[r] / region_n[r], reverse=True)
    return {
        "regions": [[r, round(region_sum[r] / region_n[r] / 10, 2), region_n[r]] for r in regions],
        "subscores": [[k, round(sub_sum[k] / sub_n[k] / 10, 2)] for k in SUBSCORES if sub_n[k]],
        "national": round(sum(region_sum.values()) / sum(region_n.values()) / 10, 2),
        "diffuses": diffuses, "total": total,
    }


def load_data():
    """Charge les agrégats. Si le xlsx source est présent, recalcule et
    rafraîchit le cache ; sinon on lit le cache commité (reproductible offline)."""
    if SRC.exists():
        data = extract_from_xlsx()
        CACHE.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        return data, "xlsx (cache rafraîchi)"
    if CACHE.exists():
        return json.loads(CACHE.read_text(encoding="utf-8")), "cache offline"
    raise SystemExit(
        f"Aucune source : ni {SRC} ni le cache {CACHE}. "
        "Placez le xlsx e-Satis ou committez viz/data_satisfaction.json.")


DATA, ORIGIN = load_data()

# KPI / besoins / charts : structure HTML standard, rendu par script custom.
# Couverture des besoins (badges) conservée : B8 imposé + contexte dimensions.
besoins = [
    {"code": "B8", "label": "Taux global de satisfaction par région — 2020", "status": "ok"},
    {"code": "Ctx", "label": "Détail par dimension de l'expérience patient", "status": "ctx"},
]
slicers = [
    {"dim": "campagne", "label": "Campagne", "type": "tiles",
     "options": [["2020", "2020 · e-Satis 48h MCO"]]},
    {"dim": "tri", "label": "Trier les régions", "type": "tiles",
     "options": [["desc", "Meilleures"], ["asc", "À améliorer"]]},
]
# icon : emoji affiché dans la pastille .kpi-ic (champ optionnel du nouveau style)
kpis = [
    {"id": "k_nat", "label": "Satisfaction nationale", "color": "#118dff",
     "note": "moyenne /10", "icon": "★"},
    {"id": "k_best", "label": "Région n°1", "color": "#1a9e57",
     "note": "", "icon": "\U0001F3C6"},
    {"id": "k_worst", "label": "Région à améliorer", "color": "#e8590c",
     "note": "", "icon": "⚠"},
    {"id": "k_nreg", "label": "Régions évaluées", "color": "#7048e8",
     "note": "campagne 2020", "icon": "\U0001F5FA"},
    {"id": "k_diff", "label": "Établissements diffusés", "color": "#e64980",
     "note": "", "icon": "\U0001F3E5"},
]
charts = [
    {"id": "c_map", "label": "Carte de la satisfaction par région", "bcode": "B8",
     "tag": "Score moyen /10 — métropole (dégradé)", "span": "col3", "tall": True},
    {"id": "c_region", "label": "Satisfaction par région (2020)", "bcode": "B8",
     "tag": "Score global ajusté moyen, sur 10 · repère = moyenne nationale",
     "span": "col3", "tall": True},
    {"id": "c_sub", "label": "Satisfaction par dimension", "bcode": "Ctx",
     "tag": "Score moyen national par axe de l'expérience patient", "span": "col3", "tall": True},
    {"id": "c_nb", "label": "Établissements évalués par région", "bcode": "Ctx",
     "tag": "Nombre d'établissements diffusés (score publiable)", "span": "col3", "tall": True},
]

CUSTOM = r"""
const D = window.__SATIS__;
if(window.__FRGEO__){echarts.registerMap('france', window.__FRGEO__);}
const charts = {};
['c_map','c_region','c_sub','c_nb'].forEach(id=>charts[id]=echarts.init(document.getElementById(id),'chu'));
let tri = 'desc';

// dégradé par valeur cohérent avec le moteur générique : faible #cfe1f7 -> fort #118dff
const lerp=(a,b,t)=>Math.round(a+(b-a)*t);
function ramp(t){const u=Math.max(0,Math.min(1,t));
  return `rgb(${lerp(0xcf,0x11,u)},${lerp(0xe1,0x8d,u)},${lerp(0xf7,0xff,u)})`;}
const score=v=>(+v).toFixed(2)+' /10';

function render(){
  const regs = [...D.regions];                 // [region, note/10, nb_etab] triées desc
  const best = regs[0], worst = regs[regs.length-1];

  // ---- carte choroplèthe : score moyen /10 par région (métropole) ----
  const scores=regs.map(d=>d[1]);
  charts.c_map.setOption({backgroundColor:'transparent',
    tooltip:Object.assign({},TTI,{formatter:p=>p.value!=null?`${p.name}: <b>${(+p.value).toFixed(2)}</b> /10`:`${p.name}: n/d`}),
    visualMap:{type:'continuous',min:Math.floor(Math.min(...scores)*10)/10,
      max:Math.ceil(Math.max(...scores)*10)/10,calculable:true,left:8,bottom:8,
      itemWidth:12,itemHeight:120,
      inRange:{color:['#f3c0c0','#fde9a9','#bfe6b6','#1a9e57']},textStyle:{color:'#605e5c',fontSize:10}},
    series:[{type:'map',map:'france',roam:false,
      data:regs.map(d=>({name:d[0],value:d[1]})),
      itemStyle:{borderColor:'#fff',borderWidth:1,areaColor:'#f0f1f4'},
      emphasis:{label:{show:false},itemStyle:{areaColor:'#fab005'}}}]},true);

  // ---- KPI (cartes en élévation, icônes, pastille delta sur la région n°1) ----
  document.getElementById('k_nat').querySelector('.val').textContent = D.national.toFixed(2);
  document.getElementById('k_best').querySelector('.val').textContent = best[0];
  document.getElementById('k_best').querySelector('.note').textContent = score(best[1]);
  document.getElementById('k_worst').querySelector('.val').textContent = worst[0];
  document.getElementById('k_worst').querySelector('.note').textContent = score(worst[1]);
  document.getElementById('k_nreg').querySelector('.val').textContent = regs.length;
  document.getElementById('k_diff').querySelector('.val').textContent = fmt(D.diffuses);
  document.getElementById('k_diff').querySelector('.note').textContent =
    'sur '+fmt(D.total)+' ('+Math.round(100*D.diffuses/D.total)+'%)';

  // pastille delta : écart de la région n°1 vs moyenne nationale (réutilise .kpi-delta)
  setDelta('k_best', best[1]-D.national);
  setDelta('k_worst', worst[1]-D.national);

  document.getElementById('insight').innerHTML =
    `Satisfaction nationale <b>${D.national.toFixed(2)}/10</b> en 2020 (${fmt(D.diffuses)} établissements diffusés). `
    +`Meilleure région : <b>${best[0]}</b> (${best[1].toFixed(2)}) ; à améliorer : <b>${worst[0]}</b> (${worst[1].toFixed(2)}).`;

  // #scope : rappel du périmètre actif (campagne + tri courant)
  const sc=document.getElementById('scope');
  if(sc)sc.innerHTML='<span class="dot"></span>Campagne 2020 · '+(tri==='asc'?'tri : à améliorer':'tri : meilleures');

  // ---- B8 : régions, barres horizontales (dégradé par valeur) + repère national ----
  const ordered = tri==='asc' ? [...regs].reverse() : [...regs];
  const forBar = [...ordered].reverse();       // ECharts y-cat de bas en haut
  const rmin=Math.min(...forBar.map(d=>d[1])), rmax=Math.max(...forBar.map(d=>d[1]));
  charts.c_region.setOption({backgroundColor:'transparent',grid:{...GRID,left:'3%',right:'9%'},
    tooltip:Object.assign({},TT,{valueFormatter:v=>v.toFixed(2)+' /10'}),
    xAxis:{type:'value',max:10,...AX,...SPL},
    yAxis:{type:'category',data:forBar.map(d=>d[0]),...AX,
      axisLabel:{color:'#605e5c',width:170,overflow:'truncate'}},
    series:[{type:'bar',barWidth:'62%',
      label:{show:true,position:'right',color:'#605e5c',fontSize:10,fontWeight:600,
        formatter:o=>(+o.value).toFixed(2)},
      data:forBar.map(d=>({value:d[1],
        itemStyle:{color:d[0]===ordered[0][0]?'#0b5cad':ramp(rmax===rmin?1:(d[1]-rmin)/(rmax-rmin)),
          borderRadius:[0,5,5,0]}})),
      markLine:{silent:true,symbol:'none',lineStyle:{color:'#e8590c',type:'dashed'},
        data:[{xAxis:D.national,label:{formatter:'Nat. '+D.national.toFixed(2),
          color:'#e8590c',position:'insideEndTop'}}]}}]},true);

  // ---- dimensions : barres horizontales + data label de valeur ----
  const sub=[...D.subscores].sort((a,b)=>a[1]-b[1]);
  const smin=Math.min(...sub.map(d=>d[1])), smax=Math.max(...sub.map(d=>d[1]));
  charts.c_sub.setOption({backgroundColor:'transparent',grid:{...GRID,left:'3%',right:'9%'},
    tooltip:Object.assign({},TT,{valueFormatter:v=>v.toFixed(2)+' /10'}),
    xAxis:{type:'value',max:10,...AX,...SPL},
    yAxis:{type:'category',data:sub.map(d=>d[0]),...AX,axisLabel:{color:'#605e5c',width:150,overflow:'truncate'}},
    series:[{type:'bar',barWidth:'62%',
      label:{show:true,position:'right',color:'#605e5c',fontSize:10,fontWeight:600,
        formatter:o=>(+o.value).toFixed(2)},
      data:sub.map(d=>({value:d[1],
        itemStyle:{color:ramp(smax===smin?1:(d[1]-smin)/(smax-smin)),borderRadius:[0,5,5,0]}}))}]},true);

  // ---- nb établissements : barres horizontales + data label de valeur fr-FR ----
  const nb=[...regs].sort((a,b)=>a[2]-b[2]);
  const nmax=Math.max(...nb.map(d=>d[2]));
  charts.c_nb.setOption({backgroundColor:'transparent',grid:{...GRID,left:'3%',right:'9%'},
    tooltip:Object.assign({},TT,{}),
    xAxis:{type:'value',...AX,...SPL},
    yAxis:{type:'category',data:nb.map(d=>d[0]),...AX,axisLabel:{color:'#605e5c',width:150,overflow:'truncate',fontSize:10}},
    series:[{type:'bar',barWidth:'62%',
      label:{show:true,position:'right',color:'#605e5c',fontSize:10,fontWeight:600,
        formatter:o=>fmt(o.value)},
      data:nb.map(d=>({value:d[2],
        itemStyle:{color:ramp(nmax?d[2]/nmax:1),borderRadius:[0,5,5,0]}}))}]},true);
}

// alimente la pastille .kpi-delta (verte si >=0, rouge sinon) en points vs national
function setDelta(id, diff){
  const el=document.getElementById(id); if(!el)return;
  const badge=el.querySelector('.kpi-delta'); if(!badge)return;
  const d=Math.round(diff*10)/10;
  badge.style.display='';
  badge.className='kpi-delta '+(d>0?'up':d<0?'down':'flat');
  badge.textContent=(d>0?'▲ +':d<0?'▼ ':'= ')+d.toFixed(1)+' pt';
}

// le helper de slicer ajoute un bouton 'Tous' non pertinent ici -> on fixe l'état initial
function initTiles(dim,val){document.querySelectorAll('#sl_'+dim+' button').forEach(b=>
  b.setAttribute('aria-pressed', String(b.dataset.v===val)));}
initTiles('tri','desc'); initTiles('campagne','2020');
// slicer "tri" (tiles) ; "campagne" verrouillée sur 2020
document.querySelectorAll('#sl_tri button').forEach(b=>b.onclick=()=>{
  if(b.dataset.v==='all')return;
  document.querySelectorAll('#sl_tri button').forEach(x=>x.setAttribute('aria-pressed','false'));
  b.setAttribute('aria-pressed','true'); tri=b.dataset.v; render();});
// chips : pas de cross-filter ici (mono-campagne, mesure = moyenne)
document.getElementById('chips').innerHTML =
  '<span class="chips-empty">Campagne 2020 · score global ajusté e-Satis 48h MCO</span>';
const ro=new ResizeObserver(es=>{for(const e of es){const i=echarts.getInstanceByDom(e.target);if(i)i.resize();}});
Object.values(charts).forEach(c=>ro.observe(c.getDom()));
window.addEventListener('resize',()=>Object.values(charts).forEach(c=>c.resize()));
render();
"""

# pré-sélectionner les tiles par défaut (2020 + tri desc) côté HTML : on s'appuie
# sur le rendu standard, puis le script force l'état via aria-pressed au 1er render.
custom = f"window.__SATIS__ = {json.dumps(DATA, ensure_ascii=False)};\n" + CUSTOM
GEOJSON = (ROOT / "scripts/viz/assets/fr_regions.geojson").read_text(encoding="utf-8")

html = dc.page(
    title="Satisfaction", sub="Tableau de bord décisionnel",
    src="e-Satis 48h MCO 2020 (open data) · agrégats", active="satisfaction",
    besoins=besoins, slicers=slicers, kpis=kpis, charts=charts,
    custom_script=custom, geojson=GEOJSON,
    foot="Score global ajusté converti sur 10. Établissements sous le seuil de diffusion exclus "
         "(cf. profiling P3). Carte = métropole. Prototype — sera reconstruit dans Power BI / Tableau sur Hive.",
)
OUT.write_text(html, encoding="utf-8")
print(f"Écrit : {OUT} ({len(html)//1024} Ko · {len(html)} octets) · source {ORIGIN} · "
      f"national {DATA['national']}/10 · {len(DATA['regions'])} régions · "
      f"{DATA['diffuses']}/{DATA['total']} diffusés")
